"""
PP-OCRv6 Local Studio — Professional OCR Workbench
启动: python3 webapp/server.py
访问: http://localhost:8765
"""
import json, time, io, base64, sqlite3, uuid, os, csv, html
from collections import Counter
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import datetime
from urllib.parse import quote
import math
import numpy as np
import yaml
from PIL import Image, ImageDraw
import onnxruntime as ort
from fastapi import FastAPI, File, UploadFile, HTTPException, Query, Body
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, PlainTextResponse, Response
from fastapi.staticfiles import StaticFiles
import uvicorn

try:
    import cv2
except Exception:
    cv2 = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# ── 路径 ─────────────────────────────────────────────────────────────────────
ROOT      = Path(__file__).parent.parent
ONNX      = ROOT / "ppocrv6_onnx"
MODEL_VARIANTS = {
    "tiny": {
        "label": "PP-OCRv6 Tiny",
        "path": ROOT / "ppocrv6_onnx",
        "official_size": "1.5M",
        "use_case": "浏览器 / 极轻端侧",
    },
    "small": {
        "label": "PP-OCRv6 Small",
        "path": ROOT / "ppocrv6_small_onnx",
        "official_size": "7.7M",
        "use_case": "移动端 / 本地应用",
    },
    "medium": {
        "label": "PP-OCRv6 Medium",
        "path": ROOT / "ppocrv6_medium_onnx",
        "official_size": "34.5M",
        "use_case": "服务器 / 高精度本地",
    },
}
STATIC    = Path(__file__).parent / "static"
DATA      = Path(__file__).parent / "data"
UPLOADS   = DATA / "uploads"
ANNOTATED = DATA / "annotated"
THUMBS    = DATA / "thumbs"
DB_PATH   = DATA / "history.db"
CFG_PATH  = DATA / "config.json"

for d in (UPLOADS, ANNOTATED, THUMBS):
    d.mkdir(parents=True, exist_ok=True)

# ── 默认配置 ─────────────────────────────────────────────────────────────────
DEFAULT_CFG = {
    "model_variant": "tiny",     # tiny | small | medium
    "det_thresh": 0.20,
    "box_thresh": 0.40,
    "unclip":     1.40,
    "max_edge":   960,
    "min_size":    3,
    "provider":   "auto",        # auto | coreml | cpu
    "save_history": True,
    "thread_count": 4,
}

def load_config():
    if CFG_PATH.exists():
        try:
            return {**DEFAULT_CFG, **json.loads(CFG_PATH.read_text())}
        except Exception:
            pass
    return dict(DEFAULT_CFG)

def save_config(cfg: dict):
    merged = {**DEFAULT_CFG, **cfg}
    CFG_PATH.write_text(json.dumps(merged, indent=2))
    return merged

# ── 模型参数（固定） ────────────────────────────────────────────────────────
DET_MEAN  = np.array([0.485, 0.456, 0.406], dtype=np.float32)
DET_STD   = np.array([0.229, 0.224, 0.225], dtype=np.float32)
REC_H     = 48
REC_MAX_W = 2400

# ── 模型单例 ─────────────────────────────────────────────────────────────────
_det_sess = _rec_sess = _char_list = _current_provider = _current_model_variant = None


def available_model_variants():
    variants = {}
    for key, meta in MODEL_VARIANTS.items():
        root = meta["path"]
        ok = (root / "det" / "inference.onnx").exists() and (root / "rec" / "inference.onnx").exists()
        variants[key] = {
            "key": key,
            "label": meta["label"],
            "official_size": meta["official_size"],
            "use_case": meta["use_case"],
            "installed": ok,
        }
    return variants


def get_model_root(model_variant: str | None = None):
    key = model_variant or load_config().get("model_variant", "tiny")
    if key not in MODEL_VARIANTS:
        key = "tiny"
    root = MODEL_VARIANTS[key]["path"]
    if not ((root / "det" / "inference.onnx").exists() and (root / "rec" / "inference.onnx").exists()):
        key = "tiny"
        root = MODEL_VARIANTS[key]["path"]
    return key, root

def get_models(force_provider: str = None, force_model: str = None):
    global _det_sess, _rec_sess, _char_list, _current_provider, _current_model_variant
    cfg = load_config()
    want = force_provider or cfg.get("provider", "auto")
    model_key, model_root = get_model_root(force_model or cfg.get("model_variant", "tiny"))

    if _det_sess is not None and _current_provider == want and _current_model_variant == model_key:
        return _det_sess, _rec_sess, _char_list

    avail = ort.get_available_providers()
    prefer = []
    if want == "coreml" and "CoreMLExecutionProvider" in avail:
        prefer = ["CoreMLExecutionProvider", "CPUExecutionProvider"]
    elif want == "cpu":
        prefer = ["CPUExecutionProvider"]
    else:  # auto
        if "CoreMLExecutionProvider" in avail:
            prefer.append("CoreMLExecutionProvider")
        prefer.append("CPUExecutionProvider")

    opts = ort.SessionOptions()
    opts.inter_op_num_threads = int(cfg.get("thread_count", 4))
    opts.intra_op_num_threads = int(cfg.get("thread_count", 4))

    _det_sess = ort.InferenceSession(str(model_root/"det"/"inference.onnx"),
                                     sess_options=opts, providers=prefer)
    _rec_sess = ort.InferenceSession(str(model_root/"rec"/"inference.onnx"),
                                     sess_options=opts, providers=prefer)
    char_json = model_root / "rec" / "char_dict.json"
    if char_json.exists():
        with open(char_json, encoding="utf-8") as f:
            d = json.load(f)
    else:
        rec_cfg = yaml.safe_load((model_root / "rec" / "inference.yml").read_text(encoding="utf-8"))
        d = rec_cfg.get("PostProcess", {}).get("character_dict", [])
        if not d:
            raise RuntimeError(f"Cannot load OCR character dictionary from {model_root / 'rec'}")
    _char_list = [""] + d + [" "]
    _current_provider = want
    _current_model_variant = model_key
    return _det_sess, _rec_sess, _char_list

def current_backend():
    if _det_sess is None:
        return "未加载"
    return _det_sess.get_providers()[0]

def current_model_variant():
    return _current_model_variant or load_config().get("model_variant", "tiny")

def content_disposition(filename: str) -> str:
    """RFC 5987 filename header with an ASCII fallback for Chinese filenames."""
    clean = (filename or "ocr").replace("\n", "_").replace("\r", "_").replace('"', "_")
    fallback = "".join(ch if 32 <= ord(ch) < 127 else "_" for ch in clean) or "ocr"
    return f'attachment; filename="{fallback}"; filename*=UTF-8\'\'{quote(clean)}'

# ── 数据库 ───────────────────────────────────────────────────────────────────
def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with db() as c:
        c.execute("""CREATE TABLE IF NOT EXISTS history(
            id          TEXT PRIMARY KEY,
            created_at  REAL NOT NULL,
            filename    TEXT NOT NULL,
            n_boxes     INT,
            det_ms      INT,
            rec_ms      INT,
            total_ms    INT,
            backend     TEXT,
            text        TEXT,
            lines_json  TEXT,
            tables_json TEXT,
            upload_path TEXT,
            thumb_path  TEXT,
            annotated_path TEXT
        )""")
        # Migrate older schema (add tables_json if missing)
        cols = [r[1] for r in c.execute("PRAGMA table_info(history)").fetchall()]
        if "tables_json" not in cols:
            c.execute("ALTER TABLE history ADD COLUMN tables_json TEXT")
        c.commit()

# ── 推理逻辑 ─────────────────────────────────────────────────────────────────
def det_preprocess(img, max_edge):
    w, h = img.size
    r = min(1.0, max_edge / max(w, h))
    nw = max(32, round(w * r / 32) * 32)
    nh = max(32, round(h * r / 32) * 32)
    arr = np.array(img.resize((nw, nh), Image.LANCZOS).convert("RGB"),
                   dtype=np.float32) / 255.0
    arr = (arr - DET_MEAN) / DET_STD
    return arr.transpose(2, 0, 1)[np.newaxis], w/nw, h/nh, nw, nh

def flood_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size):
    binary   = (prob > det_thresh).astype(np.uint8).ravel()
    labeled  = np.zeros(pw * ph, dtype=np.int32)
    cur = 0
    boxes = []
    for start in range(pw * ph):
        if binary[start] != 1 or labeled[start] != 0:
            continue
        cur += 1
        stack = [start]
        labeled[start] = cur
        xs, ys, vals = [], [], []
        while stack:
            p = stack.pop()
            x, y = p % pw, p // pw
            xs.append(x); ys.append(y); vals.append(prob[y, x])
            for nx2, ny2 in ((x-1,y),(x+1,y),(x,y-1),(x,y+1)):
                if 0 <= nx2 < pw and 0 <= ny2 < ph:
                    idx = ny2 * pw + nx2
                    if binary[idx] == 1 and labeled[idx] == 0:
                        labeled[idx] = cur
                        stack.append(idx)
        bw = max(xs)-min(xs)+1; bh = max(ys)-min(ys)+1
        if min(bw, bh) < min_size: continue
        if float(np.mean(vals)) < box_thresh: continue
        d = bw * bh * unclip / (2 * (bw + bh))
        boxes.append(dict(
            x0=max(0,(min(xs)-d)*sx), y0=max(0,(min(ys)-d)*sy),
            x1=(max(xs)+d)*sx,        y1=(max(ys)+d)*sy,
            cy=(min(ys)+max(ys))/2*sy
        ))
    boxes.sort(key=lambda b: (round(b["cy"]/10)*10, b["x0"]))
    return boxes


def order_points_clockwise(pts):
    pts = np.asarray(pts, dtype=np.float32)
    if pts.shape != (4, 2):
        return pts

    x_span = float(pts[:, 0].max() - pts[:, 0].min())
    y_span = float(pts[:, 1].max() - pts[:, 1].min())
    ordered = np.zeros((4, 2), dtype=np.float32)

    # For long OCR text boxes, the classic sum/diff ordering can collapse when
    # a slanted box touches an image edge. Split the box into left/right (or
    # top/bottom for vertical text) pairs instead; this preserves all corners.
    if x_span >= y_span:
        by_x = pts[np.argsort(pts[:, 0])]
        left = by_x[:2][np.argsort(by_x[:2, 1])]
        right = by_x[2:][np.argsort(by_x[2:, 1])]
        ordered[0], ordered[3] = left[0], left[1]
        ordered[1], ordered[2] = right[0], right[1]
    else:
        by_y = pts[np.argsort(pts[:, 1])]
        top = by_y[:2][np.argsort(by_y[:2, 0])]
        bottom = by_y[2:][np.argsort(by_y[2:, 0])]
        ordered[0], ordered[1] = top[0], top[1]
        ordered[3], ordered[2] = bottom[0], bottom[1]

    return ordered


def contour_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size):
    """DB-style rotated boxes from the probability map.

    The first local version used axis-aligned connected components. That is fast,
    but it breaks badly on perspective/rotated text such as business cards. This
    contour path keeps the text angle and lets recognition crop with perspective
    rectification, closer to PaddleOCR's official pipeline.
    """
    if cv2 is None:
        return flood_boxes(prob, pw, ph, sx, sy, det_thresh, box_thresh, unclip, min_size)

    bitmap = (prob > det_thresh).astype(np.uint8)
    kernel = np.ones((2, 2), dtype=np.uint8)
    bitmap = cv2.dilate(bitmap, kernel, iterations=1)
    contours, _ = cv2.findContours(bitmap * 255, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    boxes = []
    for contour in contours:
        if contour.shape[0] < 3:
            continue
        rect = cv2.minAreaRect(contour)
        (cx, cy), (w, h), _ = rect
        if min(w, h) < min_size:
            continue

        pts = cv2.boxPoints(rect).astype(np.float32)
        mask = np.zeros(prob.shape, dtype=np.uint8)
        cv2.fillPoly(mask, [pts.astype(np.int32)], 1)
        score = float(cv2.mean(prob, mask=mask)[0])
        if score < box_thresh:
            continue

        raw_poly = np.column_stack([pts[:, 0] * sx, pts[:, 1] * sy]).astype(float)
        ordered = order_points_clockwise(raw_poly)
        edge = ordered[1] - ordered[0]
        angle = math.degrees(math.atan2(float(edge[1]), float(edge[0])))
        while angle <= -90:
            angle += 180
        while angle > 90:
            angle -= 180

        center = pts.mean(axis=0, keepdims=True)
        # Approximate DB unclip for a rectangle. Horizontal text usually needs
        # much less padding than slanted perspective text; over-expanding it can
        # pull in background texture and hurt recognition.
        base_scale = max(1.0, min(float(unclip), 2.2))
        scale = min(base_scale, 1.08) if abs(angle) < 10 else base_scale
        pts = center + (pts - center) * scale
        pts[:, 0] = np.clip(pts[:, 0], 0, pw - 1)
        pts[:, 1] = np.clip(pts[:, 1], 0, ph - 1)

        poly = np.column_stack([pts[:, 0] * sx, pts[:, 1] * sy]).astype(float)
        x0, y0 = poly.min(axis=0)
        x1, y1 = poly.max(axis=0)
        if min(x1 - x0, y1 - y0) < 2:
            continue
        boxes.append({
            "x0": float(x0), "y0": float(y0),
            "x1": float(x1), "y1": float(y1),
            "cy": float(poly[:, 1].mean()),
            "poly": poly.tolist(),
            "angle": float(angle),
            "score": score,
        })

    boxes.sort(key=lambda b: (round(b["cy"] / 10) * 10, b["x0"]))
    return boxes


def rec_tensor_from_crop(crop: Image.Image):
    if crop.width <= 1 or crop.height <= 1:
        return None
    nw = max(8, min(REC_MAX_W, round(crop.width * REC_H / crop.height)))
    arr = np.array(crop.resize((nw, REC_H), Image.LANCZOS), dtype=np.float32)/255.0
    return ((arr - 0.5) / 0.5).transpose(2,0,1)[np.newaxis]


def affine_rec_crop(img, b, pts):
    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]
    angle = float(b.get("angle", 0))
    mat = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
    cos = abs(mat[0, 0])
    sin = abs(mat[0, 1])
    dst_w = int((h * sin) + (w * cos))
    dst_h = int((h * cos) + (w * sin))
    mat[0, 2] += dst_w / 2 - w / 2
    mat[1, 2] += dst_h / 2 - h / 2
    rotated = cv2.warpAffine(arr, mat, (dst_w, dst_h), borderMode=cv2.BORDER_REPLICATE)
    rpts = np.hstack([pts, np.ones((4, 1), dtype=np.float32)]) @ mat.T
    x0, y0 = rpts.min(axis=0)
    x1, y1 = rpts.max(axis=0)
    bw = max(1.0, float(x1 - x0))
    bh = max(1.0, float(y1 - y0))
    x0 = max(0, int(x0 - bw * 0.02))
    x1 = min(dst_w, int(x1 + bw * 0.02))
    y0 = max(0, int(y0 - bh * 0.30))
    y1 = min(dst_h, int(y1 + bh * 0.30))
    if x1 <= x0 or y1 <= y0:
        return None
    crop = Image.fromarray(rotated[y0:y1, x0:x1]).convert("RGB")
    return rec_tensor_from_crop(crop)


def rec_crop(img, b):
    # Near-horizontal text is usually cleaner with a normal crop. Perspective
    # warping helps slanted business cards, but can distort large dot-matrix
    # letters that were already horizontal.
    if b.get("poly") is not None and cv2 is not None and abs(float(b.get("angle", 0))) >= 10:
        pts = order_points_clockwise(np.array(b["poly"], dtype=np.float32))
        width_a = np.linalg.norm(pts[2] - pts[3])
        width_b = np.linalg.norm(pts[1] - pts[0])
        height_a = np.linalg.norm(pts[1] - pts[2])
        height_b = np.linalg.norm(pts[0] - pts[3])
        long_side = max(width_a, width_b)
        short_side = max(1.0, max(height_a, height_b))
        if long_side / short_side >= 6:
            rotated_crop = affine_rec_crop(img, b, pts)
            if rotated_crop is not None:
                return rotated_crop
        dst_w = max(8, int(round(max(width_a, width_b))))
        dst_h = max(8, int(round(max(height_a, height_b))))
        dst = np.array([[0, 0], [dst_w - 1, 0], [dst_w - 1, dst_h - 1], [0, dst_h - 1]],
                       dtype=np.float32)
        arr = np.array(img.convert("RGB"))
        mat = cv2.getPerspectiveTransform(pts, dst)
        warped = cv2.warpPerspective(arr, mat, (dst_w, dst_h), borderMode=cv2.BORDER_REPLICATE)
        if warped.shape[0] / max(1, warped.shape[1]) >= 1.5:
            warped = np.rot90(warped)
        crop = Image.fromarray(warped).convert("RGB")
        return rec_tensor_from_crop(crop)

    x0,y0,x1,y1 = (max(0,int(b[k])) for k in ("x0","y0","x1","y1"))
    x1 = min(x1, img.width); y1 = min(y1, img.height)
    if x1<=x0 or y1<=y0: return None
    crop = img.crop((x0,y0,x1,y1)).convert("RGB")
    return rec_tensor_from_crop(crop)

def ctc_decode(logits, char_list):
    prev, out = -1, ""
    for idx in logits.argmax(axis=1):
        if idx != 0 and idx != prev:
            out += char_list[idx] if idx < len(char_list) else ""
        prev = idx
    return out

def run_ocr(img: Image.Image, cfg: dict):
    det, rec, chars = get_models(force_model=cfg.get("model_variant"))
    t0 = time.perf_counter()
    tensor, sx, sy, pw, ph = det_preprocess(img, cfg["max_edge"])
    prob = det.run(None, {det.get_inputs()[0].name: tensor})[0][0, 0]
    det_ms = (time.perf_counter() - t0) * 1000
    boxes = contour_boxes(prob, pw, ph, sx, sy,
                          cfg["det_thresh"], cfg["box_thresh"],
                          cfg["unclip"], cfg["min_size"])
    t1 = time.perf_counter()
    lines = []
    for b in boxes:
        crop = rec_crop(img, b)
        if crop is None: continue
        out = rec.run(None, {rec.get_inputs()[0].name: crop})[0][0]
        txt = ctc_decode(out, chars)
        if txt.strip():
            item = {"text": txt, "box": [b["x0"],b["y0"],b["x1"],b["y1"]]}
            if b.get("poly") is not None:
                item["poly"] = b["poly"]
            lines.append(item)
    rec_ms = (time.perf_counter() - t1) * 1000
    return lines, det_ms, rec_ms

# ── 表格识别（启发式：行Y聚类 + 列X聚类） ──────────────────────────────────
def _median(vals, default=0):
    vals = sorted(v for v in vals if v is not None)
    if not vals:
        return default
    return vals[len(vals) // 2]


def _cluster_positions(values, tol):
    groups = []
    for v in sorted(values):
        if not groups or abs(v - groups[-1][-1]) > tol:
            groups.append([v])
        else:
            groups[-1].append(v)
    return [sum(g) / len(g) for g in groups]


def detect_tables(lines: list):
    """从 OCR 文本框还原表格结构。

    这不是 PP-Structure/SLANet 那类深度表格结构模型，而是面向本地 Tiny
    OCR 的几何重建：先聚成文本行，再用 X 方向中心点聚成列。相比上一版仅
    依赖每行文字块数量的算法，这版对中文表格、缺失单元格和列数轻微变化
    更宽容。
    返回: [{rows, n_rows, n_cols, bbox, confidence}]
    """
    if len(lines) < 4:
        return []

    items = [{
        "text": str(l.get("text", "")).strip(),
        "x0": l["box"][0], "y0": l["box"][1],
        "x1": l["box"][2], "y1": l["box"][3],
        "cx": (l["box"][0] + l["box"][2]) / 2,
        "cy": (l["box"][1] + l["box"][3]) / 2,
        "w":  max(1, l["box"][2] - l["box"][0]),
        "h":  max(1, l["box"][3] - l["box"][1]),
    } for l in lines if l.get("text", "").strip() and len(l.get("box", [])) == 4]
    if len(items) < 4:
        return []

    items.sort(key=lambda x: x["cy"])
    median_h = _median([it["h"] for it in items], 12)

    # 行聚类
    rows = [[items[0]]]
    for it in items[1:]:
        row_cy = sum(x["cy"] for x in rows[-1]) / len(rows[-1])
        avg_h = (it["h"] + _median([x["h"] for x in rows[-1]], median_h)) / 2
        if abs(it["cy"] - row_cy) <= max(8, avg_h * 0.8):
            rows[-1].append(it)
        else:
            rows.append([it])
    for r in rows:
        r.sort(key=lambda x: x["x0"])

    # 找候选表格区段
    tables = []
    i = 0
    while i < len(rows):
        if len(rows[i]) < 2:
            i += 1; continue
        j = i + 1
        while j < len(rows) and len(rows[j]) >= 2:
            prev_y = sum(x["cy"] for x in rows[j-1]) / len(rows[j-1])
            cur_y = sum(x["cy"] for x in rows[j]) / len(rows[j])
            if cur_y - prev_y > median_h * 4.0:
                break
            j += 1

        run = rows[i:j]
        if len(run) >= 2:
            median_w = _median([c["w"] for r in run for c in r], 24)
            xs = [c["cx"] for r in run for c in r]
            col_centers = _cluster_positions(xs, max(18, median_w * 0.65))
            col_centers = sorted(col_centers)

            # 过密的列通常是同一单元格被切成多个短词，二次合并一下。
            if len(col_centers) > 2:
                gaps = [b - a for a, b in zip(col_centers, col_centers[1:])]
                med_gap = _median(gaps, 0)
                if med_gap:
                    col_centers = _cluster_positions(col_centers, max(18, med_gap * 0.28))

            max_cols = len(col_centers)
            aligned_rows = 0
            table_rows = []
            for r in run:
                cells = [""] * max_cols
                used = set()
                for cell in r:
                    ci = min(range(max_cols),
                             key=lambda k: abs(cell["cx"] - col_centers[k]))
                    used.add(ci)
                    cells[ci] = (cells[ci] + " " + cell["text"]).strip() \
                                if cells[ci] else cell["text"]
                if len(used) >= min(2, max_cols):
                    aligned_rows += 1
                table_rows.append(cells)

            enough_rows = len(run) >= 3 or (len(run) >= 2 and max_cols >= 3)
            confidence = aligned_rows / max(1, len(run))
            if max_cols >= 2 and enough_rows and confidence >= 0.75:
                xs2 = [c["x0"] for r in run for c in r] + [c["x1"] for r in run for c in r]
                ys2 = [c["y0"] for r in run for c in r] + [c["y1"] for r in run for c in r]
                tables.append({
                    "rows":   table_rows,
                    "n_rows": len(table_rows),
                    "n_cols": max_cols,
                    "bbox":   [min(xs2), min(ys2), max(xs2), max(ys2)],
                    "confidence": round(confidence, 3),
                })
        i = max(j, i + 1)
    return tables


def tables_to_csv(tables: list) -> str:
    buf = io.StringIO()
    buf.write("\ufeff")
    for idx, t in enumerate(tables):
        if idx > 0: buf.write("\n\n")
        w = csv.writer(buf)
        for row in t.get("rows", []):
            w.writerow(row)
    return buf.getvalue()


def tables_to_markdown(tables: list) -> str:
    parts = []
    for idx, t in enumerate(tables):
        rows = t["rows"]
        if not rows: continue
        parts.append(f"### 表格 {idx+1}\n")
        # 第一行作表头
        header = rows[0]
        parts.append("| " + " | ".join(c.replace("|","\\|") for c in header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            row = row + [""] * (len(header) - len(row))  # pad
            parts.append("| " + " | ".join(c.replace("|","\\|") for c in row[:len(header)]) + " |")
        parts.append("")
    return "\n".join(parts)


def tables_to_html(tables: list) -> str:
    parts = ['<!DOCTYPE html><html><head><meta charset="utf-8">',
             '<style>body{font-family:-apple-system,sans-serif;padding:24px;max-width:960px;margin:0 auto}',
             'h2{font-size:16px;margin:24px 0 10px;color:#333}',
             'table{border-collapse:collapse;width:100%;margin-bottom:18px;font-size:13px}',
             'th,td{border:1px solid #ddd;padding:7px 11px;text-align:left}',
             'th{background:#f5f5f7;font-weight:600}',
             'tr:nth-child(even) td{background:#fafafa}</style></head><body>']
    for idx, t in enumerate(tables):
        parts.append(f"<h2>表格 {idx+1} · {t.get('n_rows', 0)}×{t.get('n_cols', 0)}</h2>")
        parts.append("<table>")
        for ri, row in enumerate(t.get("rows", [])):
            tag = "th" if ri == 0 else "td"
            parts.append("<tr>" + "".join(
                f"<{tag}>{html.escape(str(c))}</{tag}>"
                for c in row) + "</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    return "\n".join(parts)


def tables_to_xlsx(tables: list) -> bytes:
    if Workbook is None:
        raise HTTPException(500, "openpyxl is not installed; cannot export xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="F2F4F7")
    header_font = Font(bold=True, color="111827")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not tables:
        ws = wb.create_sheet("OCR")
        ws["A1"] = ""
    for idx, t in enumerate(tables):
        ws = wb.create_sheet(f"表格{idx+1}")
        rows = t.get("rows", [])
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        for c_idx in range(1, (t.get("n_cols") or (len(rows[0]) if rows else 1)) + 1):
            letter = get_column_letter(c_idx)
            max_len = max([len(str(ws.cell(r, c_idx).value or "")) for r in range(1, ws.max_row + 1)] or [8])
            ws.column_dimensions[letter].width = min(max(10, max_len * 1.7), 42)
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def annotate_image(img: Image.Image, lines: list, save_path: Path = None):
    vis = img.convert("RGB").copy()
    draw = ImageDraw.Draw(vis)
    for item in lines:
        if item.get("poly"):
            pts = [(int(x), int(y)) for x, y in item["poly"]]
            draw.line(pts + [pts[0]], fill=(124,92,255), width=2)
        else:
            x0,y0,x1,y1 = (int(v) for v in item["box"])
            draw.rectangle([x0,y0,x1,y1], outline=(124,92,255), width=2)
    vis.thumbnail((1600, 1600), Image.LANCZOS)
    if save_path:
        vis.save(save_path, "JPEG", quality=85)
    buf = io.BytesIO()
    vis.save(buf, "JPEG", quality=82)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

def make_thumb(img: Image.Image, save_path: Path):
    thumb = img.convert("RGB").copy()
    thumb.thumbnail((160, 160), Image.LANCZOS)
    thumb.save(save_path, "JPEG", quality=78)

# ── FastAPI ──────────────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    print("Loading models...")
    get_models()
    print(f"Ready · backend: {current_backend()}")
    yield


app = FastAPI(title="PP-OCRv6 Local Studio", lifespan=lifespan)

app.mount("/static", StaticFiles(directory=str(STATIC)), name="static")
app.mount("/data",   StaticFiles(directory=str(DATA)),   name="data")

@app.get("/", response_class=HTMLResponse)
async def index():
    return (STATIC / "index.html").read_text(encoding="utf-8")

@app.get("/tutorial", response_class=HTMLResponse)
async def tutorial():
    return (STATIC / "tutorial.html").read_text(encoding="utf-8")

@app.get("/info")
async def info():
    avail = ort.get_available_providers()
    cfg = load_config()
    return {
        "backend":  current_backend(),
        "model_variant": current_model_variant(),
        "providers_available": avail,
        "models_available": available_model_variants(),
        "config":   cfg,
        "version":  "0.3.1",
        "model": {
            "name":     MODEL_VARIANTS.get(current_model_variant(), MODEL_VARIANTS["tiny"])["label"],
            "official_size": MODEL_VARIANTS.get(current_model_variant(), MODEL_VARIANTS["tiny"])["official_size"],
            "dict_size": 6906,
        }
    }

@app.get("/health")
async def health():
    return {"status": "ok", "backend": current_backend()}

@app.post("/ocr")
async def ocr_endpoint(file: UploadFile = File(...),
                       save: bool = Query(True)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Please upload an image file")
    data = await file.read()
    try:
        img = Image.open(io.BytesIO(data)).convert("RGB")
    except Exception:
        raise HTTPException(400, "Cannot decode image")
    if max(img.size) > 6000:
        img.thumbnail((4096, 4096), Image.LANCZOS)

    cfg = load_config()
    rec_id = uuid.uuid4().hex[:12]
    lines, det_ms, rec_ms = run_ocr(img, cfg)
    tables = detect_tables(lines)

    # Save assets if history enabled
    upload_path = thumb_path = annotated_path = ""
    if save and cfg.get("save_history", True):
        ext = (file.filename or "upload.jpg").split(".")[-1].lower()
        if ext not in ("jpg", "jpeg", "png", "webp", "bmp"):
            ext = "jpg"
        upload_path = f"data/uploads/{rec_id}.{ext}"
        thumb_path  = f"data/thumbs/{rec_id}.jpg"
        annotated_path = f"data/annotated/{rec_id}.jpg"
        with open(DATA / "uploads" / f"{rec_id}.{ext}", "wb") as f:
            f.write(data)
        make_thumb(img, DATA / "thumbs" / f"{rec_id}.jpg")

    annotated_b64 = annotate_image(img, lines,
        save_path=(DATA / "annotated" / f"{rec_id}.jpg") if upload_path else None)

    if save and cfg.get("save_history", True):
        with db() as c:
            c.execute("""INSERT INTO history
                (id, created_at, filename, n_boxes, det_ms, rec_ms, total_ms,
                 backend, text, lines_json, tables_json,
                 upload_path, thumb_path, annotated_path)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (rec_id, time.time(), file.filename or "untitled",
                 len(lines), round(det_ms), round(rec_ms), round(det_ms+rec_ms),
                 current_backend(),
                 "\n".join(l["text"] for l in lines),
                 json.dumps(lines, ensure_ascii=False),
                 json.dumps(tables, ensure_ascii=False),
                 upload_path, thumb_path, annotated_path))
            c.commit()

    return JSONResponse({
        "id": rec_id,
        "filename": file.filename,
        "lines":    lines,
        "tables":   tables,
        "text":     "\n".join(l["text"] for l in lines),
        "det_ms":   round(det_ms),
        "rec_ms":   round(rec_ms),
        "total_ms": round(det_ms + rec_ms),
        "n_boxes":  len(lines),
        "n_tables": len(tables),
        "backend":  current_backend(),
        "model_variant": current_model_variant(),
        "annotated": annotated_b64,
        "thumb_url": "/" + thumb_path if thumb_path else None,
    })

# ── History ──────────────────────────────────────────────────────────────────
@app.get("/history")
async def history_list(limit: int = 50, offset: int = 0, q: str = ""):
    with db() as c:
        if q:
            rows = c.execute(
                """SELECT id,created_at,filename,n_boxes,total_ms,backend,thumb_path,tables_json
                   FROM history WHERE filename LIKE ? OR text LIKE ?
                   ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                (f"%{q}%", f"%{q}%", limit, offset)).fetchall()
            total = c.execute(
                "SELECT COUNT(*) FROM history WHERE filename LIKE ? OR text LIKE ?",
                (f"%{q}%", f"%{q}%")).fetchone()[0]
        else:
            rows = c.execute(
                """SELECT id,created_at,filename,n_boxes,total_ms,backend,thumb_path,tables_json
                   FROM history ORDER BY created_at DESC LIMIT ? OFFSET ?""",
                (limit, offset)).fetchall()
            total = c.execute("SELECT COUNT(*) FROM history").fetchone()[0]
    items = []
    for r in rows:
        item = dict(r)
        try:
            item["n_tables"] = len(json.loads(item.pop("tables_json") or "[]"))
        except Exception:
            item["n_tables"] = 0
        items.append(item)
    return {"total": total, "items": items}

@app.get("/history/{rec_id}")
async def history_detail(rec_id: str):
    with db() as c:
        row = c.execute("SELECT * FROM history WHERE id=?", (rec_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Not found")
        d = dict(row)
        d["lines"]  = json.loads(d.pop("lines_json")  or "[]")
        d["tables"] = json.loads(d.pop("tables_json") or "[]")
        return d


@app.get("/history/{rec_id}/export")
async def history_export(rec_id: str, fmt: str = Query("csv", pattern="^(csv|xlsx|md|html|json|txt)$")):
    with db() as c:
        row = c.execute("SELECT * FROM history WHERE id=?", (rec_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Not found")
    tables = json.loads(row["tables_json"] or "[]")
    text   = row["text"] or ""
    filename = (row["filename"] or "ocr").rsplit(".", 1)[0]

    if fmt == "txt":
        return PlainTextResponse(text, headers={
            "Content-Disposition": content_disposition(f"{filename}.txt")})
    if fmt == "json":
        return JSONResponse(dict(row), headers={
            "Content-Disposition": content_disposition(f"{filename}.json")})
    if fmt == "csv":
        body = tables_to_csv(tables) if tables else text
        return Response(content=body, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": content_disposition(f"{filename}.csv")})
    if fmt == "xlsx":
        if not tables:
            tables = [{"rows": [[text]], "n_rows": 1, "n_cols": 1}]
        body = tables_to_xlsx(tables)
        return Response(content=body,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": content_disposition(f"{filename}.xlsx")})
    if fmt == "md":
        body = tables_to_markdown(tables) if tables else text
        return PlainTextResponse(body, headers={
            "Content-Disposition": content_disposition(f"{filename}.md")})
    if fmt == "html":
        body = tables_to_html(tables) if tables else f"<pre>{text}</pre>"
        return HTMLResponse(body, headers={
            "Content-Disposition": content_disposition(f"{filename}.html")})

@app.delete("/history/{rec_id}")
async def history_delete(rec_id: str):
    with db() as c:
        row = c.execute("SELECT upload_path,thumb_path,annotated_path FROM history WHERE id=?",
                        (rec_id,)).fetchone()
        if row:
            for p in (row["upload_path"], row["thumb_path"], row["annotated_path"]):
                if p:
                    fp = ROOT / "webapp" / p
                    try: fp.unlink()
                    except Exception: pass
        c.execute("DELETE FROM history WHERE id=?", (rec_id,))
        c.commit()
    return {"ok": True}

@app.post("/export")
async def export_inline(payload: dict = Body(...),
                        fmt: str = Query("csv", pattern="^(csv|xlsx|md|html|txt)$")):
    """Stateless export — payload: {tables: [...], text: '...'}"""
    tables = payload.get("tables", [])
    text   = payload.get("text", "")
    name   = (payload.get("filename") or "ocr").rsplit(".",1)[0]
    if fmt == "txt":
        return PlainTextResponse(text, headers={
            "Content-Disposition": content_disposition(f"{name}.txt")})
    if fmt == "csv":
        body = tables_to_csv(tables) if tables else text
        return Response(content=body, media_type="text/csv; charset=utf-8",
                        headers={"Content-Disposition": content_disposition(f"{name}.csv")})
    if fmt == "xlsx":
        if not tables:
            tables = [{"rows": [[text]], "n_rows": 1, "n_cols": 1}]
        body = tables_to_xlsx(tables)
        return Response(content=body,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": content_disposition(f"{name}.xlsx")})
    if fmt == "md":
        body = tables_to_markdown(tables) if tables else text
        return PlainTextResponse(body, headers={
            "Content-Disposition": content_disposition(f"{name}.md")})
    if fmt == "html":
        body = tables_to_html(tables) if tables else f"<pre>{text}</pre>"
        return HTMLResponse(body, headers={
            "Content-Disposition": content_disposition(f"{name}.html")})


@app.post("/history/clear")
async def history_clear():
    with db() as c:
        c.execute("DELETE FROM history")
        c.commit()
    for d in (UPLOADS, ANNOTATED, THUMBS):
        for f in d.iterdir():
            try: f.unlink()
            except Exception: pass
    return {"ok": True}

# ── Settings ─────────────────────────────────────────────────────────────────
@app.get("/settings")
async def settings_get():
    return load_config()

@app.put("/settings")
async def settings_put(payload: dict = Body(...)):
    global _det_sess, _rec_sess, _current_provider, _current_model_variant
    cfg = save_config(payload)
    # Reload models if provider/threads changed
    _det_sess = _rec_sess = _current_provider = _current_model_variant = None
    get_models()
    return {"ok": True, "config": cfg, "backend": current_backend()}

@app.post("/settings/reset")
async def settings_reset():
    cfg = save_config(dict(DEFAULT_CFG))
    return {"ok": True, "config": cfg}

if __name__ == "__main__":
    print("PP-OCRv6 Local Studio")
    print("http://localhost:8765")
    uvicorn.run(app, host="127.0.0.1", port=8765, log_level="warning")
